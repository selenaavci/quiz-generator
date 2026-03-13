from __future__ import annotations
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = [
    "Soru Metni",
    "Seçenek Sayısı",
    "Seçenek 1",
    "Seçenek 2",
    "Seçenek 3",
    "Seçenek 4",
    "Seçenek 5",
    "Doğru Seçenek",
    "Zorluk Derecesi",
    "Departman",
    "Eğitim",
    "Konu",
    "Amaç",
    "Hazırlayan"
]

DEFAULT_PREPARER = "egitim.yonetici"
LETTER_TO_NUM = {"A": 1, "B": 2, "C": 3, "D": 4, "E": 5}
DIF_MAP = {"Kolay": 1, "Orta": 3, "Zor": 5}


@dataclass
class ExcelMeta:
    zorluk_derecesi: str = "Orta"
    departman: str = ""
    egitim: str = ""
    konu: str = ""
    amac: str = ""
    hazirlayan: str = DEFAULT_PREPARER


def _normalize_options(item: Dict[str, Any]) -> List[str]:
    opts = item.get("options")
    if not opts:
        return []

    if isinstance(opts, dict):
        keys = list(opts.keys())

        def key_rank(k: Any) -> int:
            s = str(k).strip().upper()
            if s in "ABCDE":
                return "ABCDE".index(s)
            if s.isdigit():
                return int(s) - 1
            return 999

        keys_sorted = sorted(keys, key=key_rank)
        return [str(opts[k]).strip() for k in keys_sorted][:5]

    if isinstance(opts, list):
        return [str(x).strip() for x in opts][:5]

    return []


def _correct_option_text(item: Dict[str, Any], options: List[str]) -> str:
    qtype = str(item.get("type", "")).lower().strip()

    if qtype in ("true_false", "tf"):
        return str(item.get("answer", "")).strip()

    if qtype in ("fill", "blank", "fill_blank"):
        return str(item.get("answer", "")).strip()

    corr = item.get("correct") or item.get("answer")
    if corr is None:
        return ""

    if isinstance(corr, str):
        c = corr.strip()
        if len(c) == 1 and c.upper() in "ABCDE":
            idx = "ABCDE".index(c.upper())
            return options[idx] if idx < len(options) else c.upper()
        return c

    if isinstance(corr, int):
        return options[corr] if 0 <= corr < len(options) else str(corr)

    return str(corr).strip()


def correct_to_number(item, options_list_len: int):
    """
    Returns numeric correct option index (1-based) where applicable.
    """
    qtype = str(item.get("type", "")).lower().strip()
    c = item.get("answer", None)
    if c is None:
        c = item.get("correct", None)

    if qtype in ("true_false", "tf"):
        if isinstance(c, str):
            s = c.strip().lower()
            if s in ("doğru", "dogru"):
                return 1
            if s in ("yanlış", "yanlis"):
                return 2
        if isinstance(c, int) and c in (1, 2):
            return c
        return None

    if qtype in ("fill", "blank", "fill_blank"):
        ans = item.get("answer")
        return 1 if ans else None

    if isinstance(c, str) and c.upper() in LETTER_TO_NUM:
        return LETTER_TO_NUM[c.upper()]

    if isinstance(c, str) and c.strip().isdigit():
        return int(c.strip())

    if isinstance(c, int):
        return c

    return None


def difficulty_to_points(item: Dict[str, Any], meta: ExcelMeta) -> int:
    d = item.get("difficulty") or item.get("zorluk") or meta.zorluk_derecesi

    if isinstance(d, int):
        return d

    d_str = str(d).strip()
    if d_str.isdigit():
        return int(d_str)

    return DIF_MAP.get(d_str.capitalize(), 3)


def _difficulty(item: Dict[str, Any], meta: ExcelMeta) -> int:
    return difficulty_to_points(item, meta)


def _error_row(item: Dict[str, Any], meta: ExcelMeta) -> List[Any]:
    qtype = str(item.get("type", "")).lower().strip()
    err = str(item.get("error", "")).strip()
    raw_prev = str(item.get("raw_preview", "")).strip()

    msg = f"[ERROR] {qtype}: {err}"
    if raw_prev:
        msg += f" | raw: {raw_prev[:250]}"

    return [
        msg,
        0,
        "", "", "", "", "",
        "",
        _difficulty(item, meta),
        meta.departman,
        meta.egitim,
        meta.konu,
        meta.amac,
        meta.hazirlayan,
    ]


def quiz_to_excel_rows(quiz: List[Dict[str, Any]], meta: ExcelMeta) -> List[List[Any]]:
    rows: List[List[Any]] = []

    for item in quiz:
        if isinstance(item, dict) and item.get("error"):
            rows.append(_error_row(item, meta))
            continue

        qtype = str(item.get("type", "")).lower().strip()
        soru = str(item.get("question", "")).strip()

        options = _normalize_options(item)

        if qtype in ("mcq", "multiple_choice"):
            secenek_sayisi = len(options)
            padded = (options + [""] * 5)[:5]
            opt1, opt2, opt3, opt4, opt5 = padded
            if isinstance(item, dict) and item.get("type") == "error":
                continue
            if isinstance(item, dict) and item.get("error"):
                continue

        elif qtype in ("true_false", "tf"):
            secenek_sayisi = 2
            opt1, opt2, opt3, opt4, opt5 = "Doğru", "Yanlış", "", "", ""
            if isinstance(item, dict) and item.get("type") == "error":
                continue
            if isinstance(item, dict) and item.get("error"):
                continue

        elif qtype in ("fill", "blank", "fill_blank"):
            ans = str(item.get("answer", "")).strip()
            secenek_sayisi = 1 if ans else 0
            opt1, opt2, opt3, opt4, opt5 = ans, "", "", "", ""
            if isinstance(item, dict) and item.get("type") == "error":
                continue
            if isinstance(item, dict) and item.get("error"):
                continue

        else:
            secenek_sayisi = 0
            opt1 = opt2 = opt3 = opt4 = opt5 = ""

        dogru = correct_to_number(
            item,
            len(options) if qtype in ("mcq", "multiple_choice") else secenek_sayisi
        )

        zorluk = _difficulty(item, meta)

        row = [
            soru,
            secenek_sayisi,
            opt1,
            opt2,
            opt3,
            opt4,
            opt5,
            dogru,
            zorluk,
            meta.departman,
            meta.egitim,
            meta.konu,
            meta.amac,
            meta.hazirlayan,
        ]
        rows.append(row)

    return rows


def export_quiz_to_xlsx(
    quiz: List[Dict[str, Any]],
    meta: Optional[ExcelMeta] = None,
    out_path: str = "quiz.xlsx",
    sheet_name: str = "Quiz",
    metrics: Optional[Dict[str, Any]] = None,
) -> str:
    meta = meta or ExcelMeta()

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="1A1A1A")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for r in quiz_to_excel_rows(quiz, meta):
        ws.append(r)

    widths = {
        "A": 60,
        "B": 16,
        "C": 26, "D": 26, "E": 26, "F": 26, "G": 26,
        "H": 22,
        "I": 18,
        "J": 20, "K": 22, "L": 20, "M": 20, "N": 20,
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    wrap_cols = ["A", "C", "D", "E", "F", "G", "M"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if get_column_letter(cell.column) in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    if metrics and isinstance(metrics, dict):
        mws = wb.create_sheet("Metrics")
        mws.append(["Metric", "Value"])

        mh_fill = PatternFill("solid", fgColor="1A1A1A")
        mh_font = Font(color="FFFFFF", bold=True)
        for col_idx in range(1, 3):
            cell = mws.cell(row=1, column=col_idx)
            cell.fill = mh_fill
            cell.font = mh_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for k in sorted(metrics.keys()):
            v = metrics.get(k)
            if isinstance(v, (dict, list, tuple)):
                val = str(v)
            else:
                val = "" if v is None else str(v)
            mws.append([str(k), val])

        
        mws.column_dimensions["A"].width = 44
        mws.column_dimensions["B"].width = 100
        for row in mws.iter_rows(min_row=2, max_row=mws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(out_path)
    return out_path


def export_quiz_to_csv_rows(
    quiz: List[Dict[str, Any]],
    meta: Optional[ExcelMeta] = None,
) -> Tuple[List[str], List[List[Any]]]:
    meta = meta or ExcelMeta()
    return HEADERS, quiz_to_excel_rows(quiz, meta)
